import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation # Import the tool to make dropdowns
from openpyxl.utils import get_column_letter # Get column letter for data validation
from organization.models import Course, Department

def generate_template_workbook(role_type):
    """
    Generates an openpyxl Workbook object with a specific sheet for the role_type
    and a hidden 'Data' sheet for validations.
    """
    wb = openpyxl.Workbook()
    
    # Remove default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    target_column_index = None
    validation_source_range = None
    
    # Create the Main Sheet (Students or Staff)
    if role_type == 'Student':
        ws = wb.create_sheet(title="Students")
        headers = ['First Name', 'Last Name', 'Email', 'Reg Number', 'Course']
        target_column_index = 5  # Course is 5th column

    elif role_type == 'Staff':
        ws = wb.create_sheet(title="Staff")
        headers = ['First Name', 'Last Name', 'Email', 'Employee ID', 'Department']
        target_column_index = 5  # Department is 5th column

    else:
        # Fallback
        ws = wb.create_sheet(title="Users")
        headers = ['First Name', 'Last Name', 'Email']

    ws.append(headers)

    # Create Hidden 'Data' Sheet to store a list of options
    data_sheet = wb.create_sheet(title="DataValidation")
    data_sheet.sheet_state = 'hidden'

    # Populate Data Sheet
    courses = list(Course.objects.values_list('course_name', flat=True))
    departments = list(Department.objects.values_list('department_name', flat=True))

    # Column A: Courses
    for i, course in enumerate(courses, 1):
        data_sheet.cell(row=i, column=1, value=course)
    
    # Column B: Departments
    for i, dept in enumerate(departments, 1):
        data_sheet.cell(row=i, column=2, value=dept)

    # Apply Data Validation
    if target_column_index:
        if role_type == 'Student' and courses:
            # Format: 'SheetName'!$Col$Start:$Col$End
            validation_source_range = f"'DataValidation'!$A$1:$A${len(courses)}"
        elif role_type == 'Staff' and departments:
            validation_source_range = f"'DataValidation'!$B$1:$B${len(departments)}"
        
        if validation_source_range:
            dv = DataValidation(
                type="list",
                formula1=validation_source_range,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid Selection",
                error="Please select a valid option from the dropdown list."
            )
            
            # Add the data validation to the target column
            target_letter = get_column_letter(target_column_index)
            dv.add(f"{target_letter}2:{target_letter}1000")
            
            # Add the data validation to the worksheet
            ws.add_data_validation(dv)

    return wb
